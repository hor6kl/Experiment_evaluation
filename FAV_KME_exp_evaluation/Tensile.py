from datetime import datetime
import sys
sys.path.insert(0, 'm:/ZAKAZKY/!ploTRA/')
import os
# import pickle as pck
import numpy as np
import collections
import ploTRA_JK as ploTRA
import openpyxl as xl
from openpyxl import load_workbook
from math import floor, ceil             
import TENSILE_STRENGTH_MODULUS__universal as TSM

zadani = collections.OrderedDict()
#-------------------------------------------------------------------------------
### NASTAVENI CESTY K SOUBORU SE VSTUPY ###
zadani['id_directory_main']         =  os.path.abspath(os.path.join(os.getcwd(), '..'))
zadani['id_directory_data']         =  zadani['id_directory_main']+'\\data\\'
zadani['id_directory_raw_data']     =  zadani['id_directory_main']+'\\raw_data\\'
zadani['id_directory_analyzed_exp'] =  zadani['id_directory_main']+'\\analyzed_exp\\'
if not os.path.exists(zadani['id_directory_analyzed_exp']): os.makedirs(zadani['id_directory_analyzed_exp'])
#-------------------------------------------------------------------------------

#????? NASTAVENI JMENA SOUBORU S INFORMACEMI O VZORCICH ?????
zadani['wb'] = load_workbook(filename=zadani['id_directory_data']+'experiments_Horak.xlsx', data_only=True)  #format cteni z bunky -> data_only: True=vysledek, False=vzorec
zadani['possible_group'] = zadani['wb'].sheetnames
print '------------------------------------------------------------------------------------------------------------'
print 'Possible group = ', zadani['possible_group']
#-------------------------------------------------------------------------------

#????? NAZVY SKUPIN VZORKU (NAZVY LISTU EXCELU), KTERE VYHODNOCUJEME ?????
zadani['groups'] = ['Al_TS_RT','Al_TP_RT']
zadani['name_full_graph'] = 'all'



print 'Selected group = ', zadani['groups']   
#-------------------------------------------------------------------------------
# own line style and color line
zadani['line_color_own'] = 'rbkgcmyrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkg' 
zadani['line_style_own'] = ['-','-','-','-','-','-','-','-',':',':',':',':','-.','-.','-.','-.','-','-','-','-',':',':',':',':','-.','-.','-.','-.','-','-','-','-',':',':',':',':','-.','-.','-.','-.']
# zadani['line_color_own'] = 'rgcmkrgcmkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkgrbkg' 
# zadani['line_style_own'] = ['-','-','-','-','-','',':',':',':',':','-.','-.','-.','-','-','-','-',':',':',':',':','-.','-.','-.','-.','-','-','-','-',':',':',':',':','-.','-.','-.','-.']
zadani['bool_barva_jen_do_maxima'] = bool(0)
zadani['bool_print_max_point'] = bool(1)
#-------------------------------------------------------------------------------
# nomDEF vs DEF
zadani['def_type'] = 'def'
# zadani['def_type'] = 'nomdef'
#-------------------------------------------------------------------------------
# crossarea type
zadani['type_of_cross_area'] = 'rect'
# zadani['type_of_cross_area'] = 'circ'
#-------------------------------------------------------------------------------
#????? INTERVAL PRO VYHODNOCENI ECKA ?????
zadani['bool_print_interval_in_graph'] = bool(1)
# zadani['type_of_interval'] = 'sigma_absolut'   
# zadani['interval_for_modul_sigma_absolut'] = [250e6,300e6] #absolutni velikost v Pa
zadani['type_of_interval'] = 'sigma_relativ' 
zadani['interval_for_modul_sigma_relativ'] = [0.1,0.5] #relativni velikost vuci prumernemu maximu v ramci grupy (1=100%)
# zadani['type_of_interval'] = 'epsilon' 
# zadani['interval_for_modul_epsilon'] = [0.5e-3,1.0e-3]
#-------------------------------------------------------------------------------

#????? POSUNUTI KRIVEK DO NULY ?????
zadani['orez_sily'] = [20.0]
zadani['orez_udelej'] = bool(0)
#-------------------------------------------------------------------------------

#????? SPOLECNY GRAF ?????
zadani['bool_graph_all']     = bool(1)
zadani['bool_subplot_all']   = bool(0) # 1 - subplot, 0 - pouze stress-strain
zadani['bool_graphAll_same_color_of_group'] = bool(1) # 1-krivky maji spolecnou barvu i legendu, 0-kazda krivka svoji barvu
zadani['bool_time_graph']    = bool(0)
zadani['bool_graphTime_same_color_of_group'] = bool(0) # 1-krivky maji spolecnou barvu i legendu, 0-kazda krivka svoji barvu
#????? LIMITS FOR ZOOM IN GRAPH ALL (FORCE-DISPLACEMENT)
zadani['bool_x_limits_force_disp_all']     = bool(0)
zadani['xlim_graph_force_disp_all']        = [0, 1]
zadani['bool_y_limits_force_disp_all']     = bool(0)
zadani['ylim_graph_force_disp_all']        = [0, 4]
#????? LIMITS FOR ZOOM IN GRAPH ALL (STRESS-STRAIN)
zadani['bool_x_limits_stress_strain_all']  = bool(0)
zadani['xlim_graph_stress_strain_all']     = [0,15]
zadani['bool_y_limits_stress_strain_all']  = bool(0)
zadani['ylim_graph_stress_strain_all']     = [0, 200]
#????? LIMITS FOR ZOOM IN GRAPH ALL (STRAIN-TIME)
zadani['bool_x_limits_strain_time_all'] = bool(0)
zadani['xlim_graph_strain_time_all']    = [-10, 2000]
zadani['bool_y_limits_strain_time_all'] = bool(0)
zadani['ylim_graph_strain_time_all']    = [-0.1, 2.1]
#-------------------------------------------------------------------------------

#????? GRAF PRO KAZDOU GRUPU ZVLAST ?????
zadani['bool_graph_each']        = bool(1)
zadani['bool_subplot_each']      = bool(0) # 1 - subplot, 0 - pouze stress-strain
zadani['bool_graphEach_same_color_of_group'] = bool(0) # 1-krivky maji spolecnou barvu i legendu, 0-kazda krivka svoji barvu
zadani['bool_same_range']        = bool(1) # 1-vsechny grafy stejne meritko na osach; 0-kazdy graf jine meritko
#????? LIMITS FOR ZOOM IN GRAPH ALL (FORCE-DISPLACEMENT)
zadani['bool_x_limits_force_disp_each']    = bool(0)
zadani['xlim_graph_force_disp_each']       = [0, 1]
zadani['bool_y_limits_force_disp_each']    = bool(0)
zadani['ylim_graph_force_disp_each']       = [0, 4]
#????? LIMITS FOR ZOOM IN GRAPH EACH (STRESS-STRAIN)
zadani['bool_x_limits_stress_strain_each'] = bool(0)
zadani['xlim_graph_stress_strain_each']    = [0, 8]
zadani['bool_y_limits_stress_strain_each'] = bool(0)
zadani['ylim_graph_stress_strain_each']    = [0, 400]
#-------------------------------------------------------------------------------
zadani['legend_font_size'] = 8

#????? JEDNOTKY NA GRAFECH ?????
zadani['koef_disp']    = 1000.0
zadani['label_displ']  = '$\Delta l_0$ [mm]'
zadani['koef_force']   = 1000.0
zadani['label_force']  = '$F$ [kN]'
    
zadani['koef_strain']  = 100 
zadani['label_strain'] = '$\\varepsilon$ [%]' 
zadani['koef_stress']  = 1e6 
zadani['label_stress'] = '$\sigma$ [MPa]'
zadani['label_time']   = 'time [s]'
#-------------------------------------------------------------------------------
                
#????? FORMAT SOUBORU S GRAFEM ?????
zadani['bool_show_graph_all']   = bool(0)
zadani['bool_show_graph_each']  = bool(0)
zadani['bool_format_graph_pdf'] = bool(1)
zadani['bool_format_graph_png'] = bool(1)
#-------------------------------------------------------------------------------

TSM.tensile_strength_modulus(zadani)