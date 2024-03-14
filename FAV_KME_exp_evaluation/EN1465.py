#EN 1465 - Lepidla - Stanoveni pevnosti ve smyku pri tahovem namahani preplatovanych lepenych sestav
#EN 1465 - Adhesive - Determination of tensile lap-shear strength of bonded assemblies 
from datetime import datetime

import sys
import os
import pickle as pck
import numpy as np
import collections
import ploTRA as ploTRA
import openpyxl as xl
# import xlrd
from openpyxl import load_workbook

d = collections.OrderedDict()

### NASTAVENI CESTY K SOUBORU SE VSTUPY ###
id_directory_main     =  os.path.abspath(os.path.join(os.getcwd(), '..'))
id_directory_data     =  id_directory_main+'\\data\\'
id_directory_raw_data =  id_directory_main+'\\raw_data\\'


##########################    
print '------------------------------------------------------------------------------------------------------------'

### NACTENI LISTU Z EXCELU ###
wb = load_workbook(filename=id_directory_data+'experiments_Horak.xlsx', data_only=True)  #format cteni z bunky -> data_only: True=vysledek, False=vzorec
 
possible_group = wb.sheetnames
# print 'Possible groups = ',possible_group 
    
### NAZVY SKUPIN VZORKU (NAZVY LISTU EXCELU) ###
groups = ['NOVATIT_LS_RT','NOVATIT_LS_RT_v2','NOVATIT_LS_RT_v3','NOVATIT_LS_0','NOVATIT_LS_m20']
graf_name = 'NOVATIT_LS'              

for group in groups: 
    print 'Group ID:  ',group       
    
    ### NACTENI DAT Z LISTU ###
    ws   = wb[group]
    rows = [[cell.value for cell in row] for row in ws.rows]
    cols = [[cell.value for cell in column] for column in ws.columns]
    # for row in ws.rows:
    #     for cell in row:
    #         print(cell.value)
    #     print '-------'
    
    # for column in ws.columns:
    #     for cell in column:
    #         print(cell.value)
    #     print '-------'
        
    ### NACTENI HLAVICKY ###    
    Data_start_row  = (cols[0].index('#')+1) + 1 # index je o 1 mensi nez cislo radku, a ja chci radek dalsi proto jeste +1
     
    Test_type       = ws.cell(row=cols[0].index('Test_type')+1,column=2).value  
    print 'Test type: ', Test_type
    Standard        = ws.cell(row=cols[0].index('Standard')+1,column=2).value  
    print 'Standard:  ', Standard
    Lines_to_skip   = int(ws.cell(row=cols[0].index('Lines_to_skip')+1,column=2).value)
    Material_1      = ws.cell(row=cols[0].index('Material_1')+1,column=2).value  
    print 'Material_1:  ', Material_1
    Material_2      = ws.cell(row=cols[0].index('Material_2')+1,column=2).value  
    print 'Material_2:  ', Material_2
    
    Name_of_ZS2   = ws.cell(row=cols[0].index('Name_of_ZS2')+1,column=2).value  
    print 'Name_of_ZS2:  ', Name_of_ZS2  
    Date_of_measuring   = ws.cell(row=cols[0].index('Date_of_measuring')+1,column=2).value.date()
    print 'Date_of_measuring:  ', Date_of_measuring
    print '------------------------------------------------------------------------------------------------------------'
    Quantity_TRA_row  = rows[cols[0].index('Quantity_TRA')]
    Quantity_TRA      = Quantity_TRA_row[1:Quantity_TRA_row.index(None)]
    print 'Quantity_TRA = ',Quantity_TRA
    Unit_TRA_row      = rows[cols[0].index('Unit_TRA')]
    Unit_TRA          = Unit_TRA_row[1:Unit_TRA_row.index(None)] 
    print 'Unit_TRA = ',Unit_TRA
    
    print '------------------------------------------------------------------------------------------------------------'
    
    ### NACTENI KLICU PRO DANOU GROUPU ### 
    IDs = []
    eliminated_specimens = []
    for row in range(Data_start_row,len(cols[0])+1):
        ID = ws.cell(row=row,column=1).value
        suitable = ws.cell(row=row,column=2).value
        if ID == None:
            print '*.XLSX OBSAHUJE PRAZDNE RADKY!!!!!'
        elif suitable:
            IDs.append(ID)
        else:
            eliminated_specimens.append(ID)
    print 'Specimen ID: ', IDs
    print 'Elimaneted specimens: ', eliminated_specimens
    
    ### NACTENI SEZNAMU VELICIN
    Quantity        = rows[cols[0].index('Quantity')]
    Quantity_inputs = Quantity[0:Quantity.index('#')]
    SI              = rows[cols[0].index('SI')]
    SI              = SI[0:Quantity.index('#')] 
    print 'Quantity: ',Quantity_inputs    
    ### VYTVORENI SLOVNIKOVE STRUKTURY ###
    for id in IDs:
        d[id] = collections.OrderedDict() 
        d[id]['group']                 = [group]
        d[id]['test_type']             = [Test_type] 
        d[id]['name_of_zs2']           = [Name_of_ZS2] 
        d[id]['date_of_measuring']     = [Date_of_measuring] 
        
        grip_length           = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('lj')+1).value)
        SI_grip_length        = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('lj')+1).value)
        d[id]['grip_length']  = [grip_length*SI_grip_length]
 
        ext_length            = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('l0')+1).value)
        SI_ext_length         = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('l0')+1).value)
        d[id]['ext_length']   = [ext_length*SI_ext_length]
        
        total_length          = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('L_total')+1).value)
        SI_total_length       = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('L_total')+1).value)
        d[id]['total_length'] = [total_length*SI_total_length]

        length_mat_1          = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('L_mat_1')+1).value)
        SI_length_mat_1       = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('L_mat_1')+1).value)
        d[id]['length_mat_1'] = [length_mat_1*SI_length_mat_1]
        
        length_mat_2          = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('L_mat_2')+1).value)
        SI_length_mat_2       = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('L_mat_2')+1).value)
        d[id]['length_mat_2'] = [length_mat_2*SI_length_mat_2]
        
        width_mat_1           = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('W_mat_1')+1).value)
        SI_width_mat_1        = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('W_mat_1')+1).value)
        d[id]['width_mat_1'] = [width_mat_1*SI_width_mat_1]
        
        width_mat_2           = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('W_mat_2')+1).value)
        SI_width_mat_2        = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('W_mat_2')+1).value)
        d[id]['width_mat_2'] = [width_mat_2*SI_width_mat_2]
        
        thickness_mat_1       = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('T_mat_1')+1).value)
        SI_thickness_mat_1    = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('T_mat_1')+1).value)
        d[id]['thickness_mat_1'] = [thickness_mat_1*SI_thickness_mat_1]
        
        thickness_mat_2       = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('T_mat_2')+1).value)
        SI_thickness_mat_2    = float(ws.cell(row=cols[0].index('SI')+1,column=Quantity_inputs.index('T_mat_2')+1).value)
        d[id]['thickness_mat_2'] = [thickness_mat_2*SI_thickness_mat_2]
                                       
        speed_of_crosshead = float(ws.cell(row=cols[0].index(id)+1,column=Quantity_inputs.index('v')+1).value) 
        d[id]['speed_of_crosshead']    = [speed_of_crosshead] 
        
        d[id]['extension_TRA']         = ['.TRA'] 
        d[id]['no_lines_to_skip_TRA']  = [Lines_to_skip]
        d[id]['quantity_TRA']          = Quantity_TRA
        
        data_type_TRA = []
        for i in Quantity_TRA:
            data_type_TRA.append('float')
        d[id]['data_type_TRA']         = data_type_TRA
        
        d[id]['unit_TRA']              = Unit_TRA
    print '------------------------------------------------------------------------------------------------------------'

###NACTENI DAT Z MERENI###########################################################################
keys  = d.keys()

ext,ok_code  = ploTRA.extract_data_fields_KROUPA(d,keys,['extension_TRA'])
no_lines_to_skip,ok_code  = ploTRA.extract_data_fields_KROUPA(d,keys,['no_lines_to_skip_TRA'])
data_types,ok_code  = ploTRA.extract_data_fields_KROUPA(d,keys,['data_type_TRA'])
quantity,ok_code      = ploTRA.extract_data_fields_KROUPA(d,keys,['quantity_TRA'])
unit,ok_code          = ploTRA.extract_data_fields_KROUPA(d,keys,['unit_TRA'])

print '*Finding which column is the dispalcement in .TRA files'
columns_disp,ok_code      = ploTRA.find_indices_for_column_TXT_KROUPA(quantity,'Displacement')
print '*Finding which column is the force in .TRA files'
columns_force,ok_code     = ploTRA.find_indices_for_column_TXT_KROUPA(quantity,'Standard force')

print '*Defining the data type for displacement and force using ''unit_TRA'' field in protocol'    
data_type_disp,ok_code        = ploTRA.extract_number_in_vector_fields_KROUPA(data_types,columns_disp)
data_type_force,ok_code       = ploTRA.extract_number_in_vector_fields_KROUPA(data_types,columns_force)

print '*Defining the units for displacement using ''unit_TRA'' field in protocol'    
units_disp,ok_code        = ploTRA.extract_number_in_vector_fields_KROUPA(unit,columns_disp)
print '*Defining the units for force using ''unit_TRA'' field in protocol' 
units_force,ok_code       = ploTRA.extract_number_in_vector_fields_KROUPA(unit,columns_force)

print '*Reading displacement data from .TRA files for all keys in protocol'
print 'keys: ',d.keys()
data_disp,ok_code         = ploTRA.read_TXTs_KROUPA(id_directory_raw_data,keys,ext,no_lines_to_skip,data_type_disp,columns_disp)
print '*Reading force data from .TRA files for all keys in protocol'
data_force,ok_code        = ploTRA.read_TXTs_KROUPA(id_directory_raw_data,keys,ext,no_lines_to_skip,data_type_force,columns_force)

print '*Recalculation of units of displacement data using units_TRA field in protocol' 
data_disp,ok_code         = ploTRA.arithmetic_data_fields_KROUPA(data_disp,units_disp,'*')
print '*Recalculation of units of force data using units_TRA field in protocol'
data_force,ok_code        = ploTRA.arithmetic_data_fields_KROUPA(data_force,units_force,'*')

print '*Inserting data from .TRA files into database'
for key in keys:  
    d[key],ok_code = ploTRA.create_dict_field_KROUPA(d[key],['tensile'])
    
    d[key]['tensile']['displacement'] = data_disp[key]
    d[key]['tensile']['force']        = data_force[key] 
 
###VYPOCET NEKTERYCH VELICIN#########################################################################  
print '*Inserting data from .TRA files into database (Shear_area)'
for key in keys:

#     if ploTRA.check_existence_field_KROUPA(d[key],['width_mat_1']) and ploTRA.check_existence_field_KROUPA(d[key],['width_mat_2']) and ploTRA.check_existence_field_KROUPA(d[key],['length_mat_1']) and ploTRA.check_existence_field_KROUPA(d[key],['length_mat_2']) and ploTRA.check_existence_field_KROUPA(d[key],['total_length']) :
#         print 'jsem tu?????'
#         if (not d[key]['width_mat_1']==[]) and (not d[key]['width_mat_1']==['']) and (not d[key]['width_mat_2']==[]) and (not d[key]['width_mat_2']==['']) and (not d[key]['length_mat_1']==[]) and (not d[key]['length_mat_1']==['']) and (not d[key]['length_mat_2']==[]) and (not d[key]['length_mat_2']==[''])  and (not d[key]['total_lenght']==[]) and (not d[key]['total_lenght']==['']): 
            width_shear_area  = min(d[key]['width_mat_1'][0],d[key]['width_mat_2'][0])
            length_shear_area = (d[key]['length_mat_1'][0] + d[key]['length_mat_1'][0] ) - d[key]['total_length'][0]
            d[key]['width_shear_area']  = [width_shear_area]
            d[key]['length_shear_area'] = [length_shear_area]
            d[key]['shear_area']        = [width_shear_area * length_shear_area]
            print 'shear_area:', d[key]['shear_area']
#         else:                            
#             d[key]['shear_area']  = [-1e64]
#     else:                            
#         d[key]['shear_area']  = [-1e64]
    
#MAXIMAL FORCE###
print '*Finding max force'
keys  = d.keys()  

data_displacement,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys,['tensile','displacement'])
data_force,ok_code          = ploTRA.extract_data_fields_KROUPA(d,keys,['tensile','force'])

keys_shear_area,ok_code          = ploTRA.filter_keys_field_KROUPA(d,['shear_area'])
 
keys_max_force,ok_code      = ploTRA.intersect_lists_KROUPA(data_displacement.keys(),data_force.keys())
keys_max_force,ok_code      = ploTRA.intersect_lists_KROUPA(keys_max_force,keys_shear_area)
      
for key in keys_max_force:
    d[key],ok_code   = ploTRA.create_dict_field_KROUPA(d[key],['max_force'])
     
    x,y,index,ok_code = ploTRA.find_max_point_KROUPA(data_displacement[key],data_force[key])
        
    if ok_code:
        d[key]['max_force']['displacement']  = x
        d[key]['max_force']['force']         = y
        d[key]['max_force']['index']         = index
        d[key]['max_force']['shear_stress']    = y/d[key]['shear_area'][0]  # pevnost ve smyku
 
        d[key]['tensile']['displacement_cut']  = d[key]['tensile']['displacement'][0:d[key]['max_force']['index']+1]
        d[key]['tensile']['force_cut']         = d[key]['tensile']['force'][0:d[key]['max_force']['index']+1]  
           
#AVERAGE########################################################################################
print '*Filtering keys for averaging and targets'
possible_group,ok_code      = ploTRA.find_unique_data_KROUPA(d,d.keys(),['group'])
print 'Possible groups = ',possible_group

keys_average        = collections.OrderedDict()
keys_target         = collections.OrderedDict()
    
for key_group in possible_group:
  keys,ok_code      = ploTRA.filter_keys_data_KROUPA(d,['group'],[key_group],['='])
  keys_average[key_group] = keys
            
print '*Calculating averaged values'
d,ok_code = ploTRA.create_dict_field_KROUPA(d,['average'])
for key_average in keys_average:
    d['average'],ok_code       = ploTRA.create_dict_field_KROUPA(d['average'],[key_average])
    print 'Averaging: ',key_average
           
    data_w1,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['width_mat_1'])
    data_w2,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['width_mat_2'])
    
    data_l1,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['length_mat_1'])
    data_l2,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['length_mat_2'])
    data_tl,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['total_length'])
    
    data_t1,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['thickness_mat_1'])
    data_t2,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['thickness_mat_2'])
    
    data_gl,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['grip_length'])
    data_el,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['ext_length'])
    data_As,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['shear_area'])

    data_mf,ok_code   = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['max_force','force'])
    data_mss,ok_code  = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['max_force','shear_stress'])

    data_width_shear_area,ok_code  = ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['width_shear_area'])
    data_length_shear_area,ok_code =  ploTRA.extract_data_fields_KROUPA(d,keys_average[key_average],['length_shear_area'])
     
#     print '   *w1'                                                                                    
#     w1_mean,w1_std,w1_max,w1_min,data_w1_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_w1)
#     print '   *w2'                                              
#     w2_mean,w2_std,w2_max,w2_min,data_w2_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_w2)
#     
#     print '   *l1'                                              
#     l1_mean,l1_std,l1_max,l1_min,data_l1_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_l1)
#     print '   *l2'                                              
#     l2_mean,l2_std,l2_max,l2_min,data_l2_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_l2)
#     
#     print '   *tl'                                      
#     tl_mean,tl_std,tl_max,tl_min,data_tl_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_tl)
    
    print '   *t1'                                      
    t1_mean,t1_std,t1_max,t1_min,data_t1_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_t1)
    t1_var_perc = [t1_std[0] / t1_mean[0] * 100]
    print '   *t2'                                      
    t2_mean,t2_std,t2_max,t2_min,data_t2_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_t2)
    t2_var_perc = [t2_std[0] / t2_mean[0] * 100]
    
    print '   *width_shear_area'
    wsa_mean,wsa_std,wsa_max,wsa_min,data_wsa_analyzed = ploTRA.mean_std_max_min_KROUPA(data_width_shear_area)
    wsa_var_perc = [wsa_std[0] / wsa_mean[0] * 100]
    
    print '   *length_shear_area'
    lsa_mean,lsa_std,lsa_max,lsa_min,data_lsa_analyzed = ploTRA.mean_std_max_min_KROUPA(data_length_shear_area)
    lsa_var_perc = [lsa_std[0] / lsa_mean[0] * 100]
    
#     print '   *gl'
#     gl_mean,gl_std,gl_max,gl_min,data_gl_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_gl)
#     
#     print '   *As'
#     As_mean,As_std,As_max,As_min,data_As_analyzed      = ploTRA.mean_std_max_min_KROUPA(data_As)
    
    print '   *mss'
    mss_mean,mss_std,mss_max,mss_min,data_mss_analyzed = ploTRA.mean_std_max_min_KROUPA(data_mss)
    mss_var_perc = [mss_std[0] / mss_mean[0] * 100]
    
    print '   *mf'
    mf_mean,mf_std,mf_max,mf_min,data_mf_analyzed = ploTRA.mean_std_max_min_KROUPA(data_mf)
    mf_var_perc = [mf_std[0] / mf_mean[0] * 100]

    
#     if data_w1_analyzed:
#         d['average'][key_average]['w1_mean']     = w1_mean
#         d['average'][key_average]['w1_std']      = w1_std
#         d['average'][key_average]['w1_max']      = w1_max
#         d['average'][key_average]['w1_min']      = w1_min
#         
#     if data_w2_analyzed:
#         d['average'][key_average]['w2_mean']     = w2_mean
#         d['average'][key_average]['w2_std']      = w2_std
#         d['average'][key_average]['w2_max']      = w2_max
#         d['average'][key_average]['w2_min']      = w2_min
# 
#     if data_l1_analyzed:
#         d['average'][key_average]['l1_mean']     = l1_mean
#         d['average'][key_average]['l1_std']      = l1_std
#         d['average'][key_average]['l1_var_perc'] = l1_var_perc
#         d['average'][key_average]['l1_max']      = l1_max
#         d['average'][key_average]['l1_min']      = l1_min
#         
#     if data_l2_analyzed:
#         d['average'][key_average]['l2_mean']  = l2_mean
#         d['average'][key_average]['l2_std']   = l2_std
#         d['average'][key_average]['l2_max']   = l2_max
#         d['average'][key_average]['l2_min']   = l2_min
#         
#     if data_tl_analyzed:
#         d['average'][key_average]['tl_mean']     = tl_mean
#         d['average'][key_average]['tl_std']      = tl_std
#         d['average'][key_average]['tl_max']      = tl_max
#         d['average'][key_average]['tl_min']      = tl_min

    if data_t1_analyzed:
        d['average'][key_average]['t1_mean']     = t1_mean
        d['average'][key_average]['t1_std']      = t1_std
        d['average'][key_average]['t1_var_perc'] = t1_var_perc
        d['average'][key_average]['t1_max']      = t1_max
        d['average'][key_average]['t1_min']      = t1_min
        
    if data_t2_analyzed:
        d['average'][key_average]['t2_mean']     = t2_mean
        d['average'][key_average]['t2_std']      = t2_std
        d['average'][key_average]['t2_var_perc'] = t2_var_perc
        d['average'][key_average]['t2_max']      = t2_max
        d['average'][key_average]['t2_min']      = t2_min
        
    if data_wsa_analyzed:
        d['average'][key_average]['wsa_mean']     = wsa_mean
        d['average'][key_average]['wsa_std']      = wsa_std
        d['average'][key_average]['wsa_var_perc'] = wsa_var_perc
        d['average'][key_average]['wsa_max']      = wsa_max
        d['average'][key_average]['wsa_min']      = wsa_min

    if data_lsa_analyzed:
        d['average'][key_average]['lsa_mean']     = lsa_mean
        d['average'][key_average]['lsa_std']      = lsa_std
        d['average'][key_average]['lsa_var_perc'] = lsa_var_perc
        d['average'][key_average]['lsa_max']      = lsa_max
        d['average'][key_average]['lsa_min']      = lsa_min

#     if data_gl_analyzed:
#         d['average'][key_average]['gl_mean']     = gl_mean
#         d['average'][key_average]['gl_std']      = gl_std
#         d['average'][key_average]['gl_max']      = gl_max
#         d['average'][key_average]['gl_min']      = gl_min
# 
#     if data_As_analyzed:
#         d['average'][key_average]['As_mean']  = As_mean
#         d['average'][key_average]['As_std']   = As_std
#         d['average'][key_average]['As_max']   = As_max
#         d['average'][key_average]['As_min']   = As_min

    if data_mss_analyzed:
        d['average'][key_average]['mss_mean']     = mss_mean
        d['average'][key_average]['mss_std']      = mss_std
        d['average'][key_average]['mss_var_perc'] = mss_var_perc
        d['average'][key_average]['mss_max']      = mss_max
        d['average'][key_average]['mss_min']      = mss_min
        
    if data_mf_analyzed:
        d['average'][key_average]['mf_mean']     = mf_mean
        d['average'][key_average]['mf_std']      = mf_std
        d['average'][key_average]['mf_var_perc'] = mf_var_perc
        d['average'][key_average]['mf_max']      = mf_max
        d['average'][key_average]['mf_min']      = mf_min    

#PRINT FIGURES##################################################################################
print '*Printing figure'

axis_plot_force   = []

new_path  = os.getcwd()+'\\analyzed_exp\\'
if not os.path.exists(new_path): os.makedirs(new_path)

from matplotlib import rcParams
rcParams['font.family'] = 'serif'
rcParams['font.sans-serif'] = 'Times'
rcParams['font.size'] = 8
from matplotlib.ticker import ScalarFormatter, FormatStrFormatter, MaxNLocator
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec

lw=1.0
ms=2.0

possible_color  = 'rgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmykrgbcmyk'

exp_color  = '0.8'

koef_disp   = 1000.0
koef_force  = 1000.0

sbplt_force_disp          = plt.subplot2grid((1,1),(0,0), rowspan=1, colspan=1)


keys_tensile_disp,ok_code   = ploTRA.filter_keys_field_KROUPA(d,['tensile','displacement'])
keys_tensile_force,ok_code  = ploTRA.filter_keys_field_KROUPA(d,['tensile','force'])

keys_tensile_disp_force,ok_code    = ploTRA.intersect_lists_KROUPA(keys_tensile_disp,keys_tensile_force)

for key in keys_tensile_disp_force:
    sbplt_force_disp.plot(koef_disp*d[key]['tensile']['displacement'],d[key]['tensile']['force']/koef_force,':',color=exp_color,lw=lw/4.0,ms=ms)

keys_max_disp,ok_code    = ploTRA.filter_keys_field_KROUPA(d,['max_force','displacement'])
keys_max_force,ok_code   = ploTRA.filter_keys_field_KROUPA(d,['max_force','force'])
keys_max_index,ok_code   = ploTRA.filter_keys_field_KROUPA(d,['max_force','index'])

keys_max,ok_code         = ploTRA.intersect_lists_KROUPA(keys_max_disp,keys_max_force)
keys_max,ok_code         = ploTRA.intersect_lists_KROUPA(keys_max,keys_max_index)

for key in keys_max:
    sbplt_force_disp.plot(koef_disp*d[key]['max_force']['displacement'],d[key]['max_force']['force']/koef_force,'o',color=exp_color,lw=lw,ms=ms)
    
keys_cut_disp,ok_code    = ploTRA.filter_keys_field_KROUPA(d,['tensile','displacement_cut'])
keys_cut_force,ok_code   = ploTRA.filter_keys_field_KROUPA(d,['tensile','force_cut'])

keys_cut,ok_code         = ploTRA.intersect_lists_KROUPA(keys_cut_disp,keys_cut_force)

xlim_force_disp = 0
ylim_force_disp = 0
for key in keys_cut: 
  
    for i in range(0,len(possible_group)):
        if possible_group[i] in d[key]['group']:
            id_color  = i
            
    sbplt_force_disp.plot(koef_disp*d[key]['tensile']['displacement_cut'],d[key]['tensile']['force_cut']/koef_force,'-',color=possible_color[id_color],lw=lw/2.0,ms=ms)
    
    if xlim_force_disp < max(koef_disp*d[key]['tensile']['displacement_cut']):
        xlim_force_disp = max(koef_disp*d[key]['tensile']['displacement_cut'])
    if ylim_force_disp < max(d[key]['tensile']['force_cut']/koef_force):
        ylim_force_disp = max(d[key]['tensile']['force_cut']/koef_force)                           

for i in range(0,len(possible_group)):    # kvuli jedne legende pro celou groupu
    sbplt_force_disp.plot([0],[0],'-',color=possible_color[i],lw=lw/2.0,ms=ms,label=possible_group[i])

sbplt_force_disp.grid(True)
if not axis_plot_force==[]:
    sbplt_force_disp.axis(axis_plot_force)
sbplt_force_disp.set_xlabel('$\Delta l_0$ [mm]')
sbplt_force_disp.set_ylabel('$F$ [kN]')
# sbplt_force_disp.set_xlim([0, 1.1*xlim_force_disp])
# sbplt_force_disp.set_ylim([0, 1.1*ylim_force_disp])
sbplt_force_disp.legend(loc=2, borderaxespad=0., frameon = False, prop={'size':8})
# 
# sbplt_force_disp.subplots_adjust(hspace=0.3,wspace=0.3)
# 
plt.savefig(os.getcwd()+'\\analyzed_exp\\'+graf_name+'.pdf', dpi=120, format='pdf' )
# sbplt_force_disp.savefig(os.getcwd()+'\\analyzed_exp\\'+graf_name+'.png', dpi=320, format='png' )
 
plt.clf()
plt.close('all')


#WRITE TABLES##################################################################################
print ''
tab = 35
print 'GEOMETRY DATA' 
fields_ISO  = [['width_shear_area'],['length_shear_area'],['thickness_mat_1'],['thickness_mat_2']]
units       = [                1e-3,                 1e-3,               1e-3,               1e-3]
ploTRA.write_text_table_KROUPA(d,d.keys(),fields_ISO,units,[tab])
# 
# print 'MEASUREMETS DATA' 
# fields_ISO  = [['speed_of_crosshead']]
# units       = [                     1]
# ploTRA.write_text_table_KROUPA(d,d.keys(),fields_ISO,units,[tab])
# 
print 'RESULTS FOR INDIVIDUAL SPECIMENS'
fields_ISO  = [['max_force','force'],['max_force','shear_stress']]
units       = [                    1,                         1e6]
ploTRA.write_text_table_KROUPA(d,d.keys(),fields_ISO,units,[tab])

print 'AVERAGED RESULTS'
print 'thickness_1'                            
fields_ISO  = [['t1_mean'],['t1_std'],['t1_var_perc'],['t1_min'],['t1_max']]                               
units       = [       1e-3,      1e-3,              1,      1e-3,      1e-3]                               
ploTRA.write_text_table_KROUPA(d['average'],d['average'].keys(),fields_ISO,units,[tab])

print 'thickness_2'                                
fields_ISO  = [['t2_mean'],['t2_std'],['t2_var_perc'],['t2_min'],['t2_max']]                               
units       = [       1e-3,      1e-3,              1,      1e-3,      1e-3]                               
ploTRA.write_text_table_KROUPA(d['average'],d['average'].keys(),fields_ISO,units,[tab])

print 'width_shear_area'                                
fields_ISO  = [['wsa_mean'],['wsa_std'],['wsa_var_perc'],['wsa_min'],['wsa_max']]                               
units       = [        1e-3,       1e-3,               1,       1e-3,       1e-3]                               
ploTRA.write_text_table_KROUPA(d['average'],d['average'].keys(),fields_ISO,units,[tab])

print 'length_shear_area'                                
fields_ISO  = [['lsa_mean'],['lsa_std'],['lsa_var_perc'],['lsa_min'],['lsa_max']]                               
units       = [        1e-3,       1e-3,               1,       1e-3,       1e-3]                               
ploTRA.write_text_table_KROUPA(d['average'],d['average'].keys(),fields_ISO,units,[tab])

print 'max force'                                
fields_ISO  = [['mf_mean'],['mf_std'],['mf_var_perc'],['mf_min'],['mf_max']]                               
units       = [          1,         1,              1,         1,         1]                               
ploTRA.write_text_table_KROUPA(d['average'],d['average'].keys(),fields_ISO,units,[tab])

print 'max tau'                               
fields_ISO  = [['mss_mean'],['mss_std'],['mss_var_perc'],['mss_min'],['mss_max']]                               
units       = [         1e6,        1e6,               1,        1e6,        1e6]                               
ploTRA.write_text_table_KROUPA(d['average'],d['average'].keys(),fields_ISO,units,[tab])
