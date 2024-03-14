# import load_workbook
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as pl
import os
import math

import pandas as pd

# if url.endswith('.com'):

separator = "/"  # linux "/", Windows "\\"

id_directory_python = os.path.dirname(os.path.abspath(__file__))
# print(id_directory_python)

if id_directory_python.endswith("python") or id_directory_python.endswith("Python"):
    id_directory_main = id_directory_python[:-6]
# print(id_directory_main)
id_directory_data = id_directory_main + "data"
id_directory_fig = id_directory_main + "fig"
id_directory_raw_data = id_directory_main + "raw_data"


# set file path
file_name = "experiments.xlsx"
file_name_R = 'experiments_R_curves.csv'
file_path_name = id_directory_data + "\\" + file_name
file_path_name = id_directory_data + separator + file_name
file_path_name_R = id_directory_data + separator + file_name_R
# nacte xlxs
wb = load_workbook(file_path_name)

# nacte listy
sheets_all = wb.sheetnames

print("seznam listu v xlxs: " + str(sheets_all))
print("-------------------------------------------")

# cyklus pro listy NEAKTIVNI
for list in sheets_all:
    print("prace s: " + str(list))
    sheet = wb[list]


sheet = wb.active

rows = [[cell.value for cell in row] for row in sheet.rows]
cols = [[cell.value for cell in column] for column in sheet.columns]

Quantity_TRA_row = rows[cols[0].index('Quantity_TRA')]
Quantity_TRA = Quantity_TRA_row[1:Quantity_TRA_row.index(None)]

Quantity_row = rows[cols[0].index('Quantity')]
Values_name = Quantity_row[0:Quantity_row.index(None) - 1]
print(Values_name)

# pocet sloupcu pro data
j1 = len(Values_name) + 1


SI_row = rows[cols[0].index('SI')]
SI = SI_row[0:SI_row.index(None)]
print(SI)

# najde radku s jmeny velicin
# i1 = 1
# while sheet.cell(row=i1, column=1).value != 'Quantity':
#     i1 += 1


#  najde radek s daty -1
i2 = 1
while sheet.cell(row=i2, column=1).value != '#':
    i2 += 1
Values_row = i2 + 1


# #  najde radek s nasobky velicin
# i4 = 1
# while sheet.cell(row=i4, column = 1).value != 'SI':
#     i4 += 1

# # ---------------------------------
class Data1:
    def __init__(self, data_name, data_value, data_dimension):
        range = len(data_name)
        self.data_name = data_name
        self.data_value = data_value
        self.data_dimension = data_dimension
# # ---------------------------------


B_index = Values_name.index("B")

#h0_index = Values_name.index("2h0")
hA_index = Values_name.index("2hA")
#hB_index = Values_name.index("2hB")

ai_index = Values_name.index("ai")

lj_index = Values_name.index("lj")

color_index = Values_name.index('color')
marker_index = Values_name.index('marker')
line_index = Values_name.index('line')

# cte radky
a = []
pocet = 0
while (sheet.cell(row=Values_row + pocet, column=1).value) != None:

    # nacte hodnoty
    Values = []
    for i3 in range(1, j1):
        hodnota_temp = sheet.cell(row=Values_row + pocet, column=i3).value
        Values.append(hodnota_temp)

# pole jmen a hodnot z xlsx
    a.append(Data1(Values_name, Values, SI))

    pocet += 1

# -------------------- graf


def graf_Feps(F, epsilon, name_TAR, color_set, marker_set, line_set):
    fig = pl.figure()
    ax = fig.add_subplot(111)

    for cons in range(0, len(F)):
        for i in range(0, len(epsilon[cons])):
            epsilon[cons][i] = epsilon[cons][i]*1000

    for cons in range(0, len(F)):
        pl.plot(epsilon[cons], F[cons], color=color_set[cons],
                linestyle=line_set[cons])

    # nacteni prumerovane funkce
    F_average, min_count = value_TAR_avereged(F)

    epsi_value = 0
    for cons in range(0, len(F)):
        if min_count == len(F[cons]):
            epsi_value = cons

    pl.plot(epsilon[epsi_value], F_average, color='green')

    pl.xlabel('Displacement $\delta$ [mm]')
    pl.ylabel('Force $P$ [N]')

# =============================================================================
#     name_TAR.append("DCB výsledná křivka")
# =============================================================================
    pl.legend(name_TAR)

    path_fig_png = id_directory_fig + separator + 'graf_' + list + ".png"
    path_fig_eps = id_directory_fig + separator + 'graf_' + list + ".eps"

    fig.savefig(path_fig_eps)
    fig.savefig(path_fig_png)

    print("graf_Feps vytvoren graf")

# -------------------- graf


def graf_Feps_R_curves(F, epsilon, name_TAR, color_set, marker_set, line_set):
    fig = pl.figure()
    ax = fig.add_subplot(111)

    y_average = 0
    counter = 0

    print(color_set)

    for cons in range(0, len(F)):
        for i in range(0, len(epsilon[cons])):
            epsilon[cons][i] = epsilon[cons][i]
            if epsilon[cons][i] > 60:
                y_average += F[cons][i]
                counter += 1

    for cons in range(0, len(F)):
        pl.plot(epsilon[cons], F[cons], color=color_set[cons],
                marker=marker_set[cons], linestyle=line_set[cons])

    # nacteni prumerovane funkce
    
# =============================================================================
#     mean_x_axis = [i for i in range(max(epsilon))]
# =============================================================================
    mean_x_axis = [i for i in range(30,90)]
    ys_interp = [np.interp(mean_x_axis, epsilon[i], F[i]) for i in range(len(epsilon))]
    mean_y_axis = np.mean(ys_interp, axis=0)

    pl.plot(mean_x_axis, mean_y_axis, color='green')
    
# =============================================================================
#     F_average, min_count = value_TAR_avereged(F)
# 
#     epsi_value = 0
#     for cons in range(0, len(F)):
#         if min_count == len(F[cons]):
#             epsi_value = cons
# 
#     pl.plot(epsilon[epsi_value], F_average, color='green')
# =============================================================================

    pl.xlabel('Delamination legth $a$ [mm]')
    pl.ylabel('Fracture toughness $G_I$ [J$m^{-2}$]')

# =============================================================================
#     name_TAR.append("DCB výsledná křivka")
# =============================================================================
    pl.legend(name_TAR)

    path_fig_png = id_directory_fig + "\\" + 'graf_R_' + list + ".png"
    path_fig_eps = id_directory_fig + "\\" + 'graf_R_' + list + ".eps"

    path_fig_png = id_directory_fig + separator + 'graf_R_' + list + ".png"
    path_fig_eps = id_directory_fig + separator + 'graf_R_' + list + ".eps"

    pl.plot([60, 90], [y_average/counter, y_average/counter],
            linewidth=3, color='k', linestyle='--')
    print("graf_Feps vytvoren graf")
    print(y_average/counter)

    fig.savefig(path_fig_eps)
    fig.savefig(path_fig_png)


# -------------------- TAR
def value_TAR(name_TAR):

#    path_TAR = id_directory_raw_data + "\\" + name_TAR + ".TRA"
    path_TAR = id_directory_raw_data + separator + name_TAR + ".TRA"
    f = open(path_TAR, "r", encoding='utf-8', errors='ignore')

    # for s in range(0,5):
    for s in range(0, 7):
        f.readline()

    F = []
    epsilon = []
    epsilon_n = []
    time = []

    for x in f:
        temp_values = x.split()

        F.append(temp_values[0])
        # print(F)
        epsilon.append(temp_values[2])
        # print(epsilon)
        epsilon_n.append(temp_values[2])
        #
        time.append(temp_values[1])

    for i6 in range(0, len(F)):
        epsilon[i6] = float(epsilon[i6])
        # pouze pro vykresleni v mmm
        # epsilon[i6] = float(epsilon[i6])*1000
        F[i6] = float(F[i6])

# pro grafy po jednom
        # graf_Feps(F, epsilon, name_TAR)

    print("value_TAR vytvorena data" + " " + name_TAR)

    return F, epsilon, name_TAR, epsilon_n, time
    f.close()

# ----------------------------------------


def prumer(x):

    for temp3 in range(0, len(x)):
        x_prumer += x[temp3]
    x_prumer = x_prumer/len(x)

    return x_prumer


# ----------------------------------------

def value_TAR_avereged(F):

    F_average = []
    temp1 = []
    F_prumer = 0

    for cons3 in range(0, len(F)):
        temp1.append(len(F[cons3]))
    temp1 = np.array(temp1)
    min_count = min(temp1)

    for cons2 in range(0, min_count):
        F_prumer = 0
        for cons4 in range(0, len(F)):
            F_prumer += F[cons4][cons2]
        F_prumer = F_prumer/len(F)
        F_average.append(F_prumer)

    return F_average, min_count
# ----------------------------------------




# ----------------------------------------

def vypocet_energie_DCB(F, a, epsilon_n):

    B = float(a.data_value[B_index]) * \
        float(a.data_dimension[B_index])  # sirka vzorku
    # non ahezive od osy zatizeni
    a_del = float(a.data_value[ai_index])*float(a.data_dimension[ai_index])
    delta = 0

    F_max = max(F)
    index_F_max = (np.argmax(F))

    float_lst = [float(x) for x in epsilon_n]
    deltaC = float_lst[index_F_max]  # posunuti hodnoty v bode F_max v [m]

    defor_energ = (3*F_max*deltaC)/(2*B*(a_del+delta))

    # kontorola

    return defor_energ

# ----------------------------------------


def vypocet_energie_DCB_R(F, a, epsilon_n, delam_R, delta_R):

    B = float(a.data_value[B_index]) * \
        float(a.data_dimension[B_index])  # sirka vzorku
    # non ahezive od osy zatizeni
    a_del = float(delam_R)*float(a.data_dimension[ai_index])
    delta = 0   

    deltaC = float(delta_R)

    defor_energ = (3*F*deltaC)/(2*B*(a_del+delta))

    # kontorola

    return defor_energ

# ----------------------------------------


def vypocet_energie_ENF(F, a, epsilon_n):

    B = float(a.data_value[B_index]) * \
        float(a.data_dimension[B_index])  # sirka vzorku
    # non ahezive od osy zatizeni
    a_del = float(a.data_value[ai_index])*float(a.data_dimension[ai_index])
    L = (float(a.data_value[lj_index]))*(float(a.data_dimension[lj_index]))
    # m = int(a.data_value[m_index])*a.data_dimension[m_index]
    F_max = max(F)
    index_F_max = (np.argmax(F))
    float_lst = [float(x) for x in epsilon_n]
    deltaC = float_lst[index_F_max]  # posunuti hodnoty v bode F_max v [m]

    # defor_energ = (3*m*(F_max)**2 *(a_del)**2)/(2*b_prum) #podle normy
    defor_energ = (9*a_del**2 * F_max * deltaC)/(2*B*(2*L**3 + 3*a_del**3))

    return defor_energ

# ----------------------------------------


def vypocet_energie_ENF_R(F, a, epsilon_n, delam_R, delta_R):

    B = float(a.data_value[B_index]) * \
        float(a.data_dimension[B_index])  # sirka vzorku
    # non ahezive od osy zatizeni
    a_del = float(delam_R)*float(a.data_dimension[ai_index])
    L = (float(a.data_value[lj_index]))*(float(a.data_dimension[lj_index]))
    h0 = float(a.data_value[h0_index])*(float(a.data_dimension[h0_index]))
    hA = float(a.data_value[hA_index])*(float(a.data_dimension[hA_index]))
    hB = float(a.data_value[hB_index])*(float(a.data_dimension[hB_index]))
    # m = int(a.data_value[m_index])*a.data_dimension[m_index]

    E1 = 51.10e9
    h2 = (h0+hA+hB)/3
    h = h2/2

    print(a_del)

    deltaC = float(delta_R)

    # defor_energ = (3*m*(F_max)**2 *(a_del)**2)/(2*b_prum) #podle normy
    # defor_energ = (9*a_del**2 *F *deltaC)/(2*B*(2*L**3 +3*a_del**3))  #podle Bernardina

    defor_energ = (9*F**2*a_del**2)/(16*B**2*h**3*E1)

    return defor_energ

# ----------------------------------------


def vypocet_energie_MixI_II(F, a):

    b_prum = ((float(a.data_value[5]) + float(a.data_value[6]) +
              float(a.data_value[7]))/3)*a.Values_dimension[5]  # sirka vzorku
    a_del = float(a.data_value[10])  # delamination length
    F = np.array(F)
    F_max = max(F)
    c = 0  # lever length
    L = 0  # delka mezi osami zatizeni
    Ksi = 0  # crack length correction parametr
    h_prum = ((float(a.data_value[2]) + float(a.data_value[3]) + float(
        a.data_value[4]))/3)*a.Values_dimension[2]  # specimen thisckness
    Ef = 0  # modul of elesticity in fiber direction

    GI = ((12*F ^ 2*(3*c - L) ^ 2)*(a_del + Ksi*h) ^ 2) / \
        (16*b ^ 2 * h ^ 3 * L ^ 2 * Ef)
    GII = (((9*P ^ 2*(c+L) ^ 2))*(a_del + 0.42*Ksi*h) ^ 2) / \
        (16*b ^ 2 * h ^ 3 * L ^ 2 * Ef)
    G = GI + GII

    return G


# ----------------------------------------
# ----------------------------------------

def Average(data):

    return sum(data)/len(data)


# ----------------------------------------

def Variacni_koef(F):

    var_x = 0
    avrg_F = Average(F)

    for i in F:
        var_x += (1/len(F))*(i-avrg_F)*(i-avrg_F)

    s_x = math.sqrt(var_x)

    v_x = s_x/avrg_F

    return v_x

# ----------------------------------------


def create_R_cuve(time_exp, time_R, delam_R, force, a, epsilon, delam_a0_R, time_delay_R):

    DCB_G_I = []
    delam_set = []

    print(delam_a0_R)

    if str(delam_a0_R) != 'nan':

        for i in range(0, len(time_R)):

            if str(delam_R[i]) != 'nan':

                time_del1 = float(time_R[i]) - float(time_delay_R)
                delta = float(time_exp[6]) - float(time_exp[5])
                # cyklus najde index casu v datech trhacky
                for j in range(0, len(time_exp)):
                    if float(time_del1) < float(time_exp[j]) + delta/2 and float(time_del1) > float(time_exp[j]) - delta/2:
                        break

                force_del = force[j]
                epsilon_del = epsilon[j]

                print("index " + str(j) + " force " + str(force_del) +
                      " epsilon " + str(epsilon_del) + " time " + str(time_del1))

                delam_R_compl = float(delam_a0_R) + float(delam_R[i])

                DCB_G_I.append(vypocet_energie_DCB_R(
                    force_del, a, epsilon, delam_R_compl, epsilon_del))
# =============================================================================
#                 DCB_G_I.append(vypocet_energie_ENF_R(force_del, a, epsilon, delam_R_compl, epsilon_del))
# =============================================================================
                delam_set.append(delam_R_compl)

    return DCB_G_I, delam_set

# ----------------------------------------
# ----------------------------------------


# parametry pro R-curves
R_sheet = pd.read_csv(file_path_name_R)

time_R = R_sheet.iloc[6, 7:]
delam_a0_R = R_sheet.iloc[7:, 6]
time_delay_R = R_sheet.iloc[7:, 2]

#
F = []
epsilon = []
name_TAR = []
epsilon_n = []
time = []
F_max = []
Energ_crit_array = []
cons_1 = 0


marker_set = []
line_set = []
color_set = []
DCB_G_I_set = []
delam_set_array = []

for cd in range(0, len(a)):
    if a[cd].data_value[1] == 1:

        delam_R = R_sheet.iloc[7+cd, 7:]

        F1, epsilon1, name_TAR1, epsilon_n1, time1 = value_TAR(
            a[cd].data_value[0])
        F.append(F1)
        epsilon.append(epsilon1)
        name_TAR.append(name_TAR1)
        epsilon_n.append(epsilon_n1)
        time.append(time1)
        color_set.append(a[cd].data_value[color_index])
        line_set.append(a[cd].data_value[line_index])
        marker_set.append(a[cd].data_value[marker_index])

#        F_np = np.array(F)
        F_np = F

        if "DCB" in a[cd].data_value[0]:
            Energ_crit = vypocet_energie_DCB(
                F_np[cons_1], a[cd], epsilon[cons_1])
            DCB_G_I, delam_set = create_R_cuve(
                time[cons_1], time_R, delam_R, F_np[cons_1], a[cd], epsilon[cons_1], delam_a0_R.iloc[cd], time_delay_R.iloc[cd])
        elif "ENF" in a[cd].data_value[0]:
            Energ_crit = vypocet_energie_ENF(
                F_np[cons_1], a[cd], epsilon[cons_1])
            DCB_G_I, delam_set = create_R_cuve(
                time[cons_1], time_R, delam_R, F_np[cons_1], a[cd], epsilon[cons_1], delam_a0_R.iloc[cd], time_delay_R.iloc[cd])
        elif "MIX" in a[cd].data_value[0]:
            Energ_crit = vypocet_energie_MixI_II(F_np[cons_1], a[cd])
        print("crit energie pro " + name_TAR[cons_1] + " " + str(Energ_crit))
        Energ_crit_array.append(Energ_crit)
        F_max.append(max(F_np[cons_1]))

        DCB_G_I_set.append(DCB_G_I)
        delam_set_array.append(delam_set)

        cons_1 += 1


avrg_F_max = Average(F_max)
print("Prumer maximalnich sil: " + list + " " + str(avrg_F_max))

avrg_energ_crit = Average(Energ_crit_array)
print("Prumer krit energie: " + list + " " + str(avrg_energ_crit))

var_koef_F_max = Variacni_koef(Energ_crit_array)
print("Variacni koeficient pro " + list + " " + str(var_koef_F_max))


graf_Feps(F, epsilon, name_TAR, color_set, marker_set, line_set)
graf_Feps_R_curves(DCB_G_I_set, delam_set_array, name_TAR,
                   color_set, marker_set, line_set)


print(a[1].data_value)

# F_average, temp2 = value_TAR_avereged(F)
# F_max = vypocet_energie(F_average)

# print("kriticka deformacni energie pro prumer:")
# print(F_max)


# poznamka displacement, green vysledna krivka, red MKP analyza

#


# ----------------------------------------


# ----------------------------------------
