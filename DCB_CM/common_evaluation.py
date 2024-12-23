from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd

import numpy as np

np.seterr(divide="ignore", invalid="ignore")

import os
import math

from scipy.interpolate import interp1d

import logging

# Create a logger
logger = logging.getLogger(__name__)
# Configure the logger to write to a file
logging.basicConfig(filename="Evaluation.log", encoding="utf-8", level=logging.INFO)


# --------------------
# Functions for work with dimensions data


class Common_functions:
    def __init__(self):
        pass

    def found_indexes_Voltage(self, Voltage: []) -> []:
        # Function returns list of indexes from voltage array which corresponds with voltage getting
        pass_value = 1  # corresponds with 1 V
        index_voltage = []

        first_index = None
        for counter, v in enumerate(Voltage):
            if v >= pass_value and first_index is None:
                first_index = counter
            if first_index is not None and v <= pass_value:
                index_voltage.append(first_index)
                first_index = None

        return index_voltage


# --------------------
def find_close_value_index(lst: [], x: float):
    # Find the index of a value in the list that is within a small range (range_val) of x.
    # Range value to find index is done in two iterations to consider different steps in list
    delta_in = round(len(lst) / 3)
    range_val = abs(lst[delta_in] - lst[delta_in - 1]) * 2
    # in case of fail:
    index_aprox = round(len(lst) / 2)
    for i, y in enumerate(lst):
        if abs(x - y) <= range_val:
            index_aprox = i

    range_val = abs(lst[index_aprox] - lst[index_aprox - 1])
    for i, y in enumerate(lst):
        if abs(x - y) <= range_val:
            return i
    print(f"Index was not found! Check your input.")

    return -1


def avg_range_list(X: [[float]], Y: [[float]], range_min: float) -> [float]:
    # Function computes average value on range from range_min to end of the array for list of list
    avg_value_list = [average_on_range(x, y, range_min) for (x, y) in zip(X, Y)]

    return avg_value_list


def average_on_range(X: [float], Y: [float], range_min: float) -> float:
    # Function computes average value on range from range_min to end of the array

    # finds index in array X which corresponds to value larger than range_min:
    index_min = next((index for index, x in enumerate(X) if x > range_min), None)

    average_value = np.mean(Y[index_min:])

    return average_value


# -------------------- graph


def average_curve(X: [[float]], Y: [[float]], max_min_x=None) -> []:
    # Function returns average curve from multiple curves on range from minimum_x to maximum_x on shortest curve
    # If range is specified function returns constant value which is average from max_min_x to min_max_x

    flag = False

    # Find the maximum x-value of the smallest curve
    min_max_x = min(max(x) for x in X)
    if max_min_x is None:
        max_min_x = max(min(x) for x in X)
        flag = True

    # Find the common x-scale by taking the union of all x-values which are shorten according to min_max_x
    array_x = []
    array_y = []
    for x, y in zip(X, Y):
        step = x[int(len(x) / 5)] - x[int(len(x) / 5) - 1]
        # finds index of x value which coressponds with min_max_x
        index_max = find_close_value_index(x, min_max_x)
        # finds index of x value which coressponds with min_max_x
        index_min = find_close_value_index(x, max_min_x)
        if index_min == -1:
            index_min = 0

        array_x.append(x[index_min:index_max])
        array_y.append(y[index_min:index_max])

    # creates common_x from all x values
    common_x = np.unique(np.concatenate(array_x))

    # Interpolate each curve to the common x-scale
    # This part of the code returns RuntimeWarning divede by zero  and invalid value encountered in multiply
    interpolated_curves = []
    for x, y in zip(array_x, array_y):
        f = interp1d(x, y, kind="linear", fill_value="extrapolate")
        interpolated_y = f(common_x)
        interpolated_curves.append(interpolated_y)

    # Calculate the average curve by averaging the interpolated y-values
    if flag:
        avg_y = np.mean(interpolated_curves, axis=0)
    else:
        avg_y = np.mean(interpolated_curves)
        avg_y = [avg_y, avg_y]

        common_x = [max_min_x, min_max_x]

    return common_x, avg_y


def graf_Feps(
    F, delta, name_TAR, color_set, marker_set, line_set, path_fig, indexes=None
):

    fig = plt.figure()
    ax = fig.add_subplot(111)

    #    for cons in range(0, len(F)):
    #        for i in range(0, len(delta[cons])):
    #            delta[cons][i] = delta[cons][i]*1000

    for cons in range(0, len(F)):
        plt.plot(delta[cons], F[cons], color=color_set[cons], linestyle=line_set[cons])

    if indexes is not None:
        for i in range(len(indexes)):
            for j in indexes[i]:
                plt.plot(delta[i][j], F[i][j], "x", color="black", markersize=2)

    # nacteni prumerovane funkce
    avg_x, avg_y = average_curve(delta, F)
    #    plt.plot(avg_x, avg_y, color='green')

    plt.xlabel(r"Displacement $\delta$ [m]")
    plt.ylabel(r"Force $P$ [N]")

    plt.grid()
    plt.legend(name_TAR, fontsize="6")

    path_fig_png = path_fig + ".png"
    path_fig_eps = path_fig + ".eps"

    fig.savefig(path_fig_eps)
    fig.savefig(path_fig_png, dpi=300)

    plt.close()

    print("graf_Feps vytvoren graf")


# -------------------- TAR
def value_TAR(path_TAR: str, drop_lines: int) -> tuple[list, list, list, list, list]:
    # Function reads values from .TAR file to variable column_data

    # Retrieving Name of TAR file from path_TAR
    name_TAR = path_TAR.split("/")
    name_TAR = name_TAR[len(name_TAR) - 1]
    name_TAR = name_TAR[0 : len(name_TAR) - 4]

    f = open(path_TAR, "r", encoding="utf-8", errors="ignore")

    # cycle goes through lines with informations
    for s in range(0, drop_lines - 1):
        f.readline()

    values = f.readline()
    first_row = f.readline()
    num_columns = len(first_row.split())

    column_data = []

    temp_values = first_row.split()

    # Adding the first_row of values which was used to get number of columns
    for counter, value in enumerate(temp_values):
        column_data.append([float(value)])

    for x in f:
        temp_values = x.split()

        for counter, value in enumerate(temp_values):
            column_data[counter].append(float(value))

    print("==============================")
    print(f"Data loaded for {name_TAR}")
    print(f"TAR is loaded with this column (num. {num_columns} values:")
    print(values)

    logger.info("==============================")
    logger.info(f"evaluating {name_TAR}")

    f.close()
    return column_data


def trim_data(force: [], pass_value: float):

    for counter, f in enumerate(force):
        if f > pass_value:
            return counter
    return -1


def trim_displacement(displ: []):

    new_displ = []
    init_value = displ[0]

    for counter, d in enumerate(displ):
        new_displ.append(d - init_value)

    return new_displ


# ----------------------------------------


def Variacni_koef(F) -> float:

    var_x = 0
    avrg_F = np.average(F)

    for i in F:
        var_x += (1 / len(F)) * (i - avrg_F) * (i - avrg_F)

    s_x = math.sqrt(var_x)

    v_x = s_x / avrg_F

    return v_x


# ----------------------------------------
class Dimensions_data:
    def __init__(self, data_name, data_value, data_dimension):
        range = len(data_name)
        self.data_name = data_name
        self.data_value = data_value
        self.data_dimension = data_dimension


def import_excel_values(id_directory_python: str) -> ["Dimensions_data"]:
    # Function imports values from experiment excel

    if id_directory_python.endswith("python") or id_directory_python.endswith("Python"):
        id_directory_main = id_directory_python[:-6]

    id_directory_data = id_directory_main + "data"

    # set file path
    file_name = "experiments.xlsx"
    file_path_name = os.path.join(id_directory_data, file_name)
    # loads xlxs file
    wb = load_workbook(file_path_name)

    # nacte listy
    sheets_all = wb.sheetnames

    print("seznam listu v xlxs: " + str(sheets_all))
    print("-------------------------------------------")

    # cyklus pro listy NEAKTIVNI
    for ex_list in sheets_all:
        print(f"Excel sheet: {str(ex_list)} is in evaluation.")
        sheet = wb[ex_list]

    sheet = wb.active

    rows = [[cell.value for cell in row] for row in sheet.rows]
    cols = [[cell.value for cell in column] for column in sheet.columns]

    Quantity_TRA_row = rows[cols[0].index("Quantity_TRA")]
    Quantity_TRA = Quantity_TRA_row[1 : Quantity_TRA_row.index(None)]

    Quantity_row = rows[cols[0].index("Quantity")]
    Values_name = Quantity_row[0 : Quantity_row.index(None) - 1]
    print(Values_name)

    # pocet sloupcu pro data
    j1 = len(Values_name) + 1

    SI_row = rows[cols[0].index("SI")]
    SI = SI_row[0 : SI_row.index(None)]
    print(SI)

    #  najde radek s daty -1
    i2 = 1
    while sheet.cell(row=i2, column=1).value != "#":
        i2 += 1
    Values_row = i2 + 1

    # cte radky
    a = []
    pocet = 0
    while (sheet.cell(row=Values_row + pocet, column=1).value) != None:

        # nacte hodnoty
        Values = []
        for i3 in range(1, j1):
            hodnota_temp = sheet.cell(row=Values_row + pocet, column=i3).value
            Values.append(hodnota_temp)

        # pole jmen a hodnot z xlsx
        a.append(Dimensions_data(Values_name, Values, SI))

        pocet += 1

    return a, ex_list


def import_delam_length_values(
    id_directory_python: str, dimensions: float
) -> ["Dimensions_data"]:

    # Reads main path
    if id_directory_python.endswith("python") or id_directory_python.endswith("Python"):
        id_directory_main = id_directory_python[:-6]

    id_directory_data = id_directory_main + "data"

    # set file path
    file_name_R = "experiments_R_curves.csv"
    file_path_name_R = os.path.join(id_directory_data, file_name_R)

    # parametry pro R-curves
    R_sheet = pd.read_csv(file_path_name_R)

    time_DL = R_sheet.iloc[6, 7:]

    delam_length_list = []
    delam_time_list = []
    strip_range_list = []

    for cd in range(7, len(R_sheet)):

        delam_R = R_sheet.iloc[cd, 7:]

        len_0_delam_R = len(delam_R)

        # cleaning list from 'nan' values
        delam_R = [
            float(x) for x in delam_R if not isinstance(x, float) or not math.isnan(x)
        ]
        time_R = time_DL[0 : len(delam_R)]

        strip_end = len_0_delam_R - len(delam_R)
        len_1_delam_R = len(delam_R)

        # cleaning list from zero values at the begining except for one
        for counter in range(0, len(delam_R) - 1):
            if not (delam_R[counter] == 0.0 and delam_R[counter + 1] == 0.0):
                break
        delam_R = delam_R[counter : len(delam_R)]
        delam_time = time_R[counter : len(time_R)]

        strip_start = len_1_delam_R - len(delam_R)
        final_length = len(delam_R)

        # returns range with number of values to delete from beggining and final length of list
        strip_range_list.append([strip_start, final_length])

        # Convering from original dimensions to new one
        delam_length = [x * dimensions for x in delam_R]
        delam_length_list.append(delam_length)

        delam_time_list.append(delam_time)

    return delam_length_list, delam_time_list, strip_range_list


class Statistical_evaluation:
    def __init__(self, X: [float]):
        self.X = X

    def Check_all(self) -> None:

        print("======================")
        print("Statistical Evaluation")

        print(f"Passed values: {self.X}")

        self.Check_for_outliers()

        B = self.B_basis_normal()

        print(f"B-bais lower: {B[0]:.4e}")
        print(f"B-bais upper: {B[1]:.4e}")

        A = self.A_basis_normal()

        print(f"A-bais lower: {A[0]:.4e}")
        print(f"A-bais upper: {A[1]:.4e}")

        print(f"standard deviation for sample: {np.std(self.X, ddof=1):.4e}")

        print(f"Average value: {np.mean(self.X):.4e}")

        print(
            f"Coefficient of variation: {100*np.std(self.X, ddof=1)/np.mean(self.X)} in [%]"
        )

    def Max_norm_residual(self) -> float:
        MNR = max((x - np.mean(self.X)) / np.std(self.X, ddof=1) for x in self.X)
        return MNR

    def Crit_value(self) -> float:
        # Function calculate Critical value
        n = len(self.X)
        alpha = 0.05
        t = 1 - alpha / 2 / n

        CV = (n - 1) / np.sqrt(n) * np.sqrt(t**2 / (n - 2 + t**2))
        return CV

    def Check_for_outliers(self):
        # Function checks for outliers (values too much above/below average
        CV = self.Crit_value()
        MNR = self.Max_norm_residual()

        if MNR < CV:
            print("MNR is smaller than CV. No outlier was detected.")
        else:
            print("MNR is larger than CV. Outlier was detected.")
            values = [abs(x - np.mean(self.X)) for x in self.X]
            max_value = max(values)
            index_max = values.index(max(values))

            print(
                f"Sample on position: {index_max} with value: {self.X[index_max]} is detected as outlier."
            )

    def B_basis_normal(self) -> [float]:
        # Function computes B-basis for the normal distribution
        # Function returns list where lower and upper B_basis is passed

        k_b = [
            None,
            20.581,
            6.157,
            4.163,
            3.408,
            3.007,
            2.756,
            2.583,
            2.454,
            2.355,
            2.276,
            2.211,
            2.156,
            2.109,
            2.069,
        ]
        B_low = np.mean(self.X) - k_b[len(self.X)] * np.std(self.X, ddof=1)
        B_up = np.mean(self.X) + k_b[len(self.X)] * np.std(self.X, ddof=1)
        return [B_low, B_up]

    def A_basis_normal(self) -> [float]:
        # Function computes A-basis for the normal distribution
        # Function returns list where lower and upper A-basis is passed

        k_a = [
            None,
            37.094,
            10.553,
            7.042,
            5.741,
            5.062,
            4.642,
            4.354,
            4.143,
            3.981,
            3.852,
            3.747,
            3.659,
            3.585,
            3.520,
        ]
        A_low = np.mean(self.X) - k_a[len(self.X)] * np.std(self.X, ddof=1)
        A_up = np.mean(self.X) + k_a[len(self.X)] * np.std(self.X, ddof=1)
        return [A_low, A_up]
